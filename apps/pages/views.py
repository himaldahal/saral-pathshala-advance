from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import Http404, HttpResponse
from django.utils import timezone
from django.core.cache import cache
from django.urls import reverse
from django.conf import settings

from .models import Course, Lecture, Enrollment, LectureProgress, Subject, Notice, CourseEnrollmentRequest, SiteSetting
from apps.cauth.models import User, StudentLevel
from apps.exam.models import Exam, ExamAttempt

# =========================================================
# HOMEPAGE & STATIC PAGES
# =========================================================

def home(request):
    """SEO-focused, content-rich homepage for Saral Pathshala. Query-cached."""
    cached_home = cache.get('homepage_data')
    if not cached_home:
        courses = list(Course.objects.filter(is_active=True).prefetch_related('subjects')[:6])
        upcoming_exams = list(Exam.objects.filter(
            is_active=True,
            start_date__gte=timezone.now()
        ).order_by('start_date')[:6])
        notices = list(Notice.objects.filter(is_active=True).order_by('-created_at')[:6])
        
        cached_home = {
            'courses': courses,
            'upcoming_exams': upcoming_exams,
            'notices': notices
        }
        cache.set('homepage_data', cached_home, 300) # Cache for 5 mins
        
    return render(request, 'home.html', {
        'courses': cached_home['courses'],
        'upcoming_exams': cached_home['upcoming_exams'],
        'notices': cached_home['notices'],
        'now': timezone.now(),
    })

def about(request):
    """About Us page."""
    return render(request, 'about.html')

def contact(request):
    """Contact Us page with dynamic course lead dropdown options."""
    courses = cache.get('contact_courses_all')
    if not courses:
        courses = list(Course.objects.filter(is_active=True))
        cache.set('contact_courses_all', courses, 600)
    return render(request, 'contact.html', {'courses': courses})

# COURSE VIEWS
def course_list(request):
    courses = cache.get('course_list_all')
    if not courses:
        courses = list(Course.objects.filter(is_active=True).prefetch_related('subjects'))
        cache.set('course_list_all', courses, 600)

    enrolled_ids = set()
    if request.user.is_authenticated:
        enrolled_ids = set(
            Enrollment.objects.filter(user=request.user)
            .values_list('course_id', flat=True)
        )

    courses_with_meta = []
    for course in courses:
        total_lectures_key = f'course_total_lectures_{course.slug}'
        total_lectures = cache.get(total_lectures_key)
        if total_lectures is None:
            total_lectures = course.total_lectures()
            cache.set(total_lectures_key, total_lectures, 600)

        total_subjects_key = f'course_total_subjects_{course.slug}'
        total_subjects = cache.get(total_subjects_key)
        if total_subjects is None:
            total_subjects = course.total_subjects()
            cache.set(total_subjects_key, total_subjects, 600)

        first_lecture_key = f'course_first_lecture_{course.slug}'
        first_lecture = cache.get(first_lecture_key)
        if first_lecture is None:
            first_lecture = course.get_first_lecture()
            cache.set(first_lecture_key, first_lecture, 600)

        courses_with_meta.append({
            'course': course,
            'is_enrolled': (course.pk in enrolled_ids) or (request.user.is_authenticated and (request.user.is_superuser or request.user.is_staff)),
            'total_lectures': total_lectures,
            'total_subjects': total_subjects,
            'first_lecture': first_lecture,
        })

    return render(request, 'courses/course_list.html', {
        'courses_with_meta': courses_with_meta,
    })


def course_detail(request, slug):
    """SEO detail page for a course displaying subjects, sections, syllabus. Cached."""
    course_key = f'course_detail_obj_{slug}'
    course = cache.get(course_key)
    if not course:
        course = get_object_or_404(Course, slug=slug, is_active=True)
        cache.set(course_key, course, 600)

    subjects_key = f'course_subjects_{slug}'
    subjects = cache.get(subjects_key)
    if not subjects:
        subjects = list(course.subjects.filter(is_active=True).prefetch_related('sections__lectures').order_by('order', 'id'))
        cache.set(subjects_key, subjects, 600)
    
    is_enrolled = False
    if request.user.is_authenticated:
        is_enrolled = request.user.is_superuser or request.user.is_staff or Enrollment.objects.filter(user=request.user, course=course).exists()
        
    return render(request, 'courses/course_detail.html', {
        'course': course,
        'subjects': subjects,
        'is_enrolled': is_enrolled,
    })


@login_required
def enroll_instantly(request, slug):
    """Creates a CourseEnrollmentRequest lead instead of instant enrollment."""
    course = get_object_or_404(Course, slug=slug, is_active=True)
    
    # Check if request already exists
    already_requested = CourseEnrollmentRequest.objects.filter(
        email=request.user.email
    ).exists()
    
    if not already_requested:
        CourseEnrollmentRequest.objects.create(
            name=request.user.full_name or request.user.email,
            phone=request.user.phone or "0000000000",
            email=request.user.email,
            course=course,
            message="Auto-generated request from logged-in student.",
            status='pending'
        )
    messages.success(request, f"Your enrollment request for '{course.name}' has been submitted successfully. A staff member will contact you soon!")
    return redirect('course_detail', slug=slug)


def enroll_request(request):
    """Guest enrollment request form. Protected by honeypot and rate-limiting."""
    if request.method == "POST":
        name = request.POST.get('name', '').strip()
        phone = request.POST.get('phone', '').strip()
        email = request.POST.get('email', '').strip()
        course_slug = request.POST.get('course_slug', '').strip()
        message = request.POST.get('message', '').strip()
        
        # Honeypot field check
        honeypot = request.POST.get('website_url', '').strip()
        if honeypot:
            messages.success(request, "Enrollment request submitted successfully.")
            return redirect('course_list')

        course = get_object_or_404(Course, slug=course_slug)
        
        if not name or not phone:
            messages.error(request, "Full name and Phone number are required.")
            return redirect('course_detail', slug=course_slug)

        # Enforce Nepalese phone basic validation matching registration
        import re
        if not re.match(r'^(97|98)\d{8}$', phone):
            messages.error(request, "Please enter a valid Nepalese phone number starting with 97 or 98 (10 digits).")
            return redirect('course_detail', slug=course_slug)

        # Lead submission rate-limiting: 30 minutes block
        cache_key = f"enroll_lead_{phone}_{course.pk}"
        if cache.get(cache_key):
            messages.warning(request, "We have already received an enrollment request from this number recently. We will contact you soon.")
            return redirect('course_detail', slug=course_slug)

        # IP-based rate-limiting to prevent lead flooding
        from apps.cauth.utils import get_client_ip
        ip_address = get_client_ip(request)
        ip_lead_key = f"enroll_lead_ip_{ip_address}"
        ip_lead_attempts = cache.get(ip_lead_key, 0)
        if ip_lead_attempts >= 5:
            messages.error(request, "Too many enrollment requests from your device. Please try again later.")
            return redirect('course_detail', slug=course_slug)

        CourseEnrollmentRequest.objects.create(
            name=name,
            phone=phone,
            email=email,
            course=course,
            message=message
        )
        cache.set(cache_key, True, 1800) # 30 min lock
        cache.set(ip_lead_key, ip_lead_attempts + 1, 3600) # 1 hour block for IP

        messages.success(request, "Thank you! Your enrollment request has been submitted. Our team will call you shortly.")
        return redirect('course_detail', slug=course_slug)
        
    return redirect('course_list')


# =========================================================
# LECTURE VIEWS
# =========================================================

def lecture_detail(request, lecture_id):
    lecture = get_object_or_404(
        Lecture.objects.select_related('section__subject'),
        pk=lecture_id
    )

    # Resolve the course from subject
    course = lecture.section.subject.courses.first()
    if not course:
        raise Http404("No course found for this lecture.")

    # Access control: force login to view any content.
    if not request.user.is_authenticated:
        messages.warning(request, "Please log in to view this lecture.")
        return redirect('login')

    # Non-preview lectures require enrollment (or staff/superuser).
    if not lecture.is_preview:
        is_enrolled = Enrollment.objects.filter(
            user=request.user, course=course
        ).exists()

        if not (request.user.is_superuser or request.user.is_staff or is_enrolled):
            messages.error(request, "You must be enrolled in this course to view this lecture.")
            return redirect('course_detail', slug=course.slug)

    # Build sidebar: subjects → sections → lectures for this course
    subjects = (
        Subject.objects
        .filter(courses=course, is_active=True)
        .prefetch_related('sections__lectures')
        .order_by('order', 'id')
    )

    # Track progress (authenticated users only)
    completed_ids = set()
    if request.user.is_authenticated:
        completed_ids = set(
            LectureProgress.objects.filter(
                user=request.user, completed=True
            ).values_list('lecture_id', flat=True)
        )

        # Mark current lecture as watched
        LectureProgress.objects.update_or_create(
            user=request.user,
            lecture=lecture,
            defaults={'completed': True}
        )

    is_enrolled = False
    if request.user.is_authenticated:
        is_enrolled = Enrollment.objects.filter(
            user=request.user, course=course
        ).exists() or request.user.is_superuser or request.user.is_staff

    next_lecture = lecture.get_next_lecture(course)
    prev_lecture = lecture.get_prev_lecture(course)

    return render(request, 'courses/lecture_detail.html', {
        'lecture': lecture,
        'course': course,
        'subjects': subjects,
        'completed_ids': completed_ids,
        'next_lecture': next_lecture,
        'prev_lecture': prev_lecture,
        'is_enrolled': is_enrolled,
    })


# =========================================================
# STUDENT DASHBOARD
# =========================================================

@login_required
def dashboard(request):
    # Enforce profile updates
    if request.method == "POST":
        if "update_profile" in request.POST:
            full_name = request.POST.get('full_name', '').strip()
            previous_institute = request.POST.get('previous_institute', '').strip()
            current_level = request.POST.get('current_level', '').strip()
            interested_course_slug = request.POST.get('interested_course', '').strip()
            
            if not full_name:
                messages.error(request, "Full name cannot be empty.")
                return redirect('dashboard')
                
            # Update profile info (except phone and date_joined)
            request.user.full_name = full_name
            request.user.previous_institute = previous_institute
            
            if current_level in [choice[0] for choice in StudentLevel.choices]:
                request.user.current_level = current_level
            elif not current_level:
                request.user.current_level = None
                
            if interested_course_slug:
                interested_course = Course.objects.filter(slug=interested_course_slug, is_active=True).first()
                request.user.interested_course = interested_course
            else:
                request.user.interested_course = None
                
            # Email Update logic (allowed ONLY if unverified)
            if not request.user.is_email_verified:
                email = request.POST.get('email', '').strip().lower()
                if email:
                    if email != request.user.email:
                        if User.objects.exclude(pk=request.user.pk).filter(email=email).exists():
                            messages.error(request, "This email is already registered to another account.")
                        else:
                            request.user.email = email
                            messages.info(request, "Your email has been updated. Please verify it to secure your account.")
                else:
                    messages.error(request, "Email address cannot be empty.")
                    return redirect('dashboard')
            
            request.user.save()
            messages.success(request, "Profile updated successfully.")
            return redirect('dashboard')
            
        elif "request_verification" in request.POST:
            # Prevent abuse using the built-in rate-limiting system
            from apps.cauth.utils import generate_and_dispatch_email_token
            success, msg = generate_and_dispatch_email_token(request.user, request)
            if success:
                messages.success(request, "Verification link sent! Please check your email inbox.")
            else:
                messages.warning(request, msg)
            return redirect('dashboard')

    enrollments = (
        Enrollment.objects
        .filter(user=request.user)
        .select_related('course')
        .order_by('-enrolled_at')
    )

    enrollment_data = []

    total_courses = enrollments.count()
    total_subjects = 0
    total_lectures = 0
    completed_lectures = 0

    recent_progress = []

    for enrollment in enrollments:
        course = enrollment.course

        lectures = (
            Lecture.objects
            .filter(section__subject__courses=course)
            .distinct()
            .order_by(
                'section__subject__order',
                'section__order',
                'order',
                'id'
            )
        )

        lecture_count = lectures.count()

        completed_count = (
            LectureProgress.objects
            .filter(
                user=request.user,
                lecture__in=lectures,
                completed=True
            )
            .count()
        )

        next_lecture = (
            lectures.exclude(
                progress__user=request.user,
                progress__completed=True
            )
            .first()
        )

        progress_percent = 0

        if lecture_count > 0:
            progress_percent = round(
                (completed_count / lecture_count) * 100
            )

        enrollment_data.append({
            'enrollment': enrollment,
            'course': course,
            'lecture_count': lecture_count,
            'completed_count': completed_count,
            'progress_percent': progress_percent,
            'subject_count': course.total_subjects(),
            'next_lecture': next_lecture,
        })

        total_subjects += course.total_subjects()
        total_lectures += lecture_count
        completed_lectures += completed_count

    overall_progress = 0

    if total_lectures > 0:
        overall_progress = round(
            (completed_lectures / total_lectures) * 100
        )

    enrolled_course_lectures = Lecture.objects.filter(
        section__subject__courses__in=[
            enrollment.course for enrollment in enrollments
        ]
    ).distinct()

    recent_progress = (
        LectureProgress.objects
        .filter(
            user=request.user,
            lecture__in=enrolled_course_lectures
        )
        .select_related(
            'lecture',
            'lecture__section',
            'lecture__section__subject'
        )
        .order_by('-watched_at')[:5]
    )

    # ── Exam Performance Analytics Chart ──────────────────────────────────────
    from django.db.models import Avg
    import json
    
    attempts = ExamAttempt.objects.filter(
        student=request.user, 
        is_submitted=True
    ).select_related('exam').order_by('completed_at')

    chart_labels = []
    chart_scores = []
    chart_averages = []

    for att in attempts:
        chart_labels.append(att.exam.title)
        total_marks = att.exam.total_marks()
        
        # Student score percentage
        score_pct = round((att.score / total_marks) * 100, 1) if total_marks > 0 else 0.0
        chart_scores.append(score_pct)
        
        # Global class average percentage for this exam
        avg_score_agg = ExamAttempt.objects.filter(
            exam=att.exam, 
            is_submitted=True
        ).aggregate(avg=Avg('score'))
        
        avg_score = avg_score_agg['avg'] or 0.0
        avg_pct = round((avg_score / total_marks) * 100, 1) if total_marks > 0 else 0.0
        chart_averages.append(avg_pct)

    context = {
        'user': request.user,
        'enrollments': enrollments,
        'enrollment_data': enrollment_data,
        'total_courses': total_courses,
        'total_subjects': total_subjects,
        'total_lectures': total_lectures,
        'completed_lectures': completed_lectures,
        'overall_progress': overall_progress,
        'recent_progress': recent_progress,
        'chart_labels_json': json.dumps(chart_labels),
        'chart_scores_json': json.dumps(chart_scores),
        'chart_averages_json': json.dumps(chart_averages),
        'has_attempts': attempts.exists(),
        'attempts': attempts,
        'active_courses': Course.objects.filter(is_active=True).order_by('name'),
        'student_levels': StudentLevel.choices,
    }

    return render(request, 'dashboard.html', context)


# =========================================================
# NOTICES (PHYSICAL EXAMS & RESULTS)
# =========================================================

def notice_list(request):
    """SEO optimized list of notices."""
    notices = Notice.objects.filter(is_active=True).order_by('-created_at')
    return render(request, 'notices/notice_list.html', {'notices': notices})


def notice_detail(request, slug):
    """SEO optimized detail view of a notice with PDF rendering."""
    notice = get_object_or_404(Notice, slug=slug, is_active=True)
    return render(request, 'notices/notice_detail.html', {'notice': notice})


# =========================================================
# DYNAMIC CACHED SITEMAP
# =========================================================

def sitemap_xml(request):
    """
    Generates a dynamic XML sitemap of all active urls.
    Cached for 6 hours to minimize DB hits.
    """
    sitemap_data = cache.get('sitemap_xml_data')
    if not sitemap_data:
        host = f"https://{request.get_host()}"
        urls = [
            f"{host}/",
            f"{host}/courses/",
            f"{host}/notices/",
            f"{host}/exams/",
        ]
        
        # Course Detail URLs
        for course in Course.objects.filter(is_active=True):
            urls.append(f"{host}{course.get_absolute_url()}")
            
        # Notice Detail URLs
        for notice in Notice.objects.filter(is_active=True):
            urls.append(f"{host}{notice.get_absolute_url()}")
            
        # Exam Detail URLs
        for exam in Exam.objects.filter(is_active=True):
            urls.append(f"{host}{reverse('exams:exam_detail', kwargs={'slug': exam.slug})}")
            
        xml = '<?xml version="1.0" encoding="UTF-8"?>\n'
        xml += '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
        for url in urls:
            xml += f'  <url>\n    <loc>{url}</loc>\n    <changefreq>daily</changefreq>\n    <priority>0.8</priority>\n  </url>\n'
        xml += '</urlset>'
        sitemap_data = xml
        cache.set('sitemap_xml_data', sitemap_data, 21600) # 6 hours cache
        
    return HttpResponse(sitemap_data, content_type='application/xml')


def robots_txt(request):
    host = f"https://{request.get_host()}"

    content = f"""User-agent: *\nAllow: /\nSitemap: {host}/sitemap.xml
                """

    return HttpResponse(content, content_type="text/plain")

def handler404(request, exception=None):
    """Custom 404 Page Not Found error handler."""
    return render(request, '404.html', status=404)


def handler500(request):
    """Custom 500 Internal Server Error handler."""
    return render(request, '500.html', status=500)


# =========================================================
# STAFF & ADMIN CUSTOM UTILITIES
# =========================================================

@login_required
def admin_import_enrollments(request):
    """
    Staff / Superuser utility to import user enrollments in bulk via CSV or JSON.
    Tracks enrolled_by and enrolled_at.
    """
    if not (request.user.is_staff or request.user.is_superuser):
        raise Http404("You do not have permission to view this page.")

    import csv
    import json

    results = None
    if request.method == "POST":
        file_type = request.POST.get('file_type')
        uploaded_file = request.FILES.get('import_file')
        
        if not uploaded_file:
            messages.error(request, "Please upload a CSV or JSON file.")
            return redirect('admin_import_enrollments')

        # Read the file content
        try:
            content = uploaded_file.read().decode('utf-8')
        except Exception as e:
            messages.error(request, f"Error reading file: {e}")
            return redirect('admin_import_enrollments')

        records = []
        errors = []
        success_count = 0

        # Parse CSV or JSON
        if file_type == 'csv' or uploaded_file.name.endswith('.csv'):
            reader = csv.DictReader(content.splitlines())
            for row_idx, row in enumerate(reader, 1):
                course_slug = row.get('course-slug') or row.get('course_slug')
                email = row.get('email')
                phone = row.get('phone') or row.get('phone_no') or row.get('phone-no')
                
                if not course_slug:
                    errors.append(f"Row {row_idx}: Missing course-slug.")
                    continue
                if not email and not phone:
                    errors.append(f"Row {row_idx}: Both email and phone are missing.")
                    continue
                records.append({
                    'course_slug': course_slug.strip(),
                    'email': email.strip() if email else None,
                    'phone': phone.strip() if phone else None,
                    'row_num': row_idx
                })
        else: # JSON
            try:
                data = json.loads(content)
                if not isinstance(data, list):
                    messages.error(request, "JSON content must be a list of objects.")
                    return redirect('admin_import_enrollments')
                for idx, item in enumerate(data, 1):
                    course_slug = item.get('course-slug') or item.get('course_slug')
                    email = item.get('email')
                    phone = item.get('phone') or item.get('phone_no') or item.get('phone-no')
                    
                    if not course_slug:
                        errors.append(f"Item {idx}: Missing course-slug.")
                        continue
                    if not email and not phone:
                        errors.append(f"Item {idx}: Both email and phone are missing.")
                        continue
                    records.append({
                        'course_slug': str(course_slug).strip(),
                        'email': str(email).strip() if email else None,
                        'phone': str(phone).strip() if phone else None,
                        'row_num': idx
                    })
            except Exception as e:
                messages.error(request, f"Invalid JSON payload: {e}")
                return redirect('admin_import_enrollments')

        # Process the records
        from django.db import transaction
        for rec in records:
            course = Course.objects.filter(slug=rec['course_slug']).first()
            if not course:
                errors.append(f"Row/Item {rec['row_num']}: Course with slug '{rec['course_slug']}' not found.")
                continue

            user = None
            # Give precedence to email
            if rec['email']:
                user = User.objects.filter(email=rec['email']).first()
            if not user and rec['phone']:
                user = User.objects.filter(phone=rec['phone']).first()

            if not user:
                errors.append(f"Row/Item {rec['row_num']}: User not found with email '{rec['email']}' or phone '{rec['phone']}'.")
                continue

            # Create enrollment
            with transaction.atomic():
                enrollment, created = Enrollment.objects.get_or_create(
                    user=user,
                    course=course,
                    defaults={
                        'enrolled_by': request.user
                    }
                )
                if created:
                    success_count += 1
                else:
                    errors.append(f"Row/Item {rec['row_num']}: Student {user.email} is already enrolled in {course.name}.")

        results = {
            'success_count': success_count,
            'errors': errors,
            'total_processed': len(records)
        }
        if success_count > 0:
            messages.success(request, f"Import complete! Enrolled {success_count} students successfully.")
        if errors:
            messages.warning(request, f"Import completed with {len(errors)} warnings/errors.")

    return render(request, 'admin_custom/import_enrollments.html', {
        'results': results
    })


@login_required
def admin_dashboard(request):
    """
    Highly customized, feature-rich admin dashboard.
    Shows real-time exam takers, scores, distributions, enrollment leads CRM, and demographics.
    Also serves as a system operations command center, supporting background queue manual flushing,
    diagnostics, student account status controls, and exam attempt clears.
    """
    if not (request.user.is_staff or request.user.is_superuser):
        raise Http404("You do not have permission to view this page.")

    from django.db.models import Avg, Max, Min
    import json
    from datetime import timedelta
    from django.core.management import call_command
    from django.core.mail import send_mail
    from apps.cauth.models import MailQueue, SMSQueue, QueueStatus
    
    # Handle Administrative Action POSTs
    if request.method == "POST":
        # 1) Portal Settings Update
        if "update_settings" in request.POST:
            site_settings = SiteSetting.objects.first()
            if not site_settings:
                site_settings = SiteSetting.objects.create()
                
            site_settings.site_name = request.POST.get('site_name', '').strip()
            site_settings.site_title = request.POST.get('site_title', '').strip()
            site_settings.site_description = request.POST.get('site_description', '').strip()
            site_settings.meta_keywords = request.POST.get('meta_keywords', '').strip()
            
            site_settings.contact_email = request.POST.get('contact_email', '').strip()
            site_settings.contact_phone = request.POST.get('contact_phone', '').strip()
            site_settings.contact_address = request.POST.get('contact_address', '').strip()
            site_settings.whatsapp_number = request.POST.get('whatsapp_number', '').strip()
            
            site_settings.social_facebook = request.POST.get('social_facebook', '').strip()
            site_settings.social_twitter = request.POST.get('social_twitter', '').strip()
            site_settings.social_instagram = request.POST.get('social_instagram', '').strip()
            site_settings.social_tiktok = request.POST.get('social_tiktok', '').strip()
            site_settings.social_linkedin = request.POST.get('social_linkedin', '').strip()
            site_settings.social_threads = request.POST.get('social_threads', '').strip()
            
            site_settings.google_analytics_id = request.POST.get('google_analytics_id', '').strip()
            
            if 'logo' in request.FILES:
                site_settings.logo = request.FILES['logo']
            if 'favicon' in request.FILES:
                site_settings.favicon = request.FILES['favicon']
                
            site_settings.save()
            messages.success(request, "Portal settings updated successfully!")
            return redirect('admin_dashboard')
            
        # 2) Process/Flush Mail Queue
        elif "process_mail_queue" in request.POST:
            try:
                call_command('process_mail_queue')
                messages.success(request, "Mail Queue processed successfully! Pending emails have been sent.")
            except Exception as e:
                messages.error(request, f"Failed to process Mail Queue: {str(e)}")
            return redirect('admin_dashboard')
            
        # 3) Process/Flush SMS Queue
        elif "process_sms_queue" in request.POST:
            try:
                call_command('process_sms_queue')
                messages.success(request, "SMS Queue processed successfully! Pending SMS messages have been sent.")
            except Exception as e:
                messages.error(request, f"Failed to process SMS Queue: {str(e)}")
            return redirect('admin_dashboard')
            
        # 3.5) Run Unified Cron Task
        elif "run_unified_cron" in request.POST:
            try:
                call_command('unified_cron')
                messages.success(request, "Unified Cron Task completed successfully! Dispatched pending OTPs, reset emails, and cleared old/used tokens.")
            except Exception as e:
                messages.error(request, f"Failed to run Unified Cron: {str(e)}")
            return redirect('admin_dashboard')
            
        # 4) Retry Failed Queues
        elif "retry_failed_queues" in request.POST:
            mail_count = MailQueue.objects.filter(status=QueueStatus.FAILED).update(status=QueueStatus.PENDING, retry_count=0)
            sms_count = SMSQueue.objects.filter(status=QueueStatus.FAILED).update(status=QueueStatus.PENDING, retry_count=0)
            messages.success(request, f"Successfully reset retry counters! Queued {mail_count} failed emails and {sms_count} failed SMS back to PENDING.")
            return redirect('admin_dashboard')

        # 4.5) Flush All Cache
        elif "flush_all_cache" in request.POST:
            try:
                cache.clear()
                messages.success(request, "All system cache cleared successfully!")
            except Exception as e:
                messages.error(request, f"Failed to clear cache: {str(e)}")
            return redirect('admin_dashboard')

        # 5) Delete a specific student attempt (reset mock exams)
        elif "delete_student_attempt" in request.POST:
            attempt_id = request.POST.get('attempt_id', '').strip()
            attempt_to_delete = get_object_or_404(ExamAttempt, pk=attempt_id)
            student = attempt_to_delete.student
            exam_title = attempt_to_delete.exam.title
            attempt_to_delete.delete()
            # Invalidate cached list for student
            cache.delete(f'exam_list_{student.pk}')
            messages.success(request, f"Mock exam attempt for '{student.get_full_name()}' on '{exam_title}' has been deleted. The student can now start a fresh attempt.")
            return redirect('admin_dashboard')

        # 6) Manually Verify Email
        elif "verify_user_email" in request.POST:
            user_id = request.POST.get('user_id', '').strip()
            user_to_verify = get_object_or_404(User, pk=user_id)
            user_to_verify.is_email_verified = True
            user_to_verify.save(update_fields=['is_email_verified'])
            messages.success(request, f"Email address for student '{user_to_verify.get_full_name()}' has been manually VERIFIED.")
            return redirect('admin_dashboard')

        # 7) Manually Verify Phone
        elif "verify_user_phone" in request.POST:
            user_id = request.POST.get('user_id', '').strip()
            user_to_verify = get_object_or_404(User, pk=user_id)
            user_to_verify.is_phone_verified = True
            user_to_verify.save(update_fields=['is_phone_verified'])
            messages.success(request, f"Phone number for student '{user_to_verify.get_full_name()}' has been manually VERIFIED.")
            return redirect('admin_dashboard')

        # 8) SMTP Diagnostics Connection Test
        elif "send_test_email" in request.POST:
            test_email = request.POST.get('test_email', '').strip()
            if not test_email:
                messages.warning(request, "Please enter a valid test email address.")
            else:
                try:
                    from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'info@saralpathshala.com')
                    send_mail(
                        subject="SMTP Diagnostic Test | Saral Pathshala Command Center",
                        message="SMTP diagnostics are working perfectly. This is a plain text fallback.",
                        from_email=from_email,
                        recipient_list=[test_email],
                        html_message="""
                        <div style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; max-width: 600px; margin: 0 auto; padding: 30px; border-radius: 16px; background-color: #ffffff; border: 1px solid #e2e8f0; box-shadow: 0 4px 12px rgba(0,0,0,0.03);">
                            <div style="text-align: center; margin-bottom: 25px;">
                                <h1 style="color: #2d89c8; margin: 0; font-size: 24px; font-weight: 700;">SMTP Server Online</h1>
                                <p style="color: #64748b; font-size: 14px; margin: 5px 0 0 0;">Saral Pathshala Custom Diagnostics</p>
                            </div>
                            <div style="padding: 20px; background-color: #f8fafc; border-radius: 12px; border: 1px dashed #cbd5e1; margin-bottom: 25px;">
                                <p style="margin: 0 0 10px 0; color: #334155; font-size: 15px; line-height: 1.6;">
                                    Hello Administrator,
                                </p>
                                <p style="margin: 0; color: #334155; font-size: 15px; line-height: 1.6;">
                                    This diagnostic mail confirms that your Django SMTP configurations inside <code>settings.py</code> are working flawlessly! Outbound transactional emails are now ready to be delivered to students.
                                </p>
                            </div>
                            <div style="border-top: 1px solid #e2e8f0; padding-top: 15px; text-align: center; color: #94a3b8; font-size: 12px;">
                                Sent from Saral Pathshala Command Center &bull; Live SMTP Check
                            </div>
                        </div>
                        """,
                        fail_silently=False,
                    )
                    messages.success(request, f"SMTP Connection Test Successful! Test email sent successfully to {test_email}.")
                except Exception as e:
                    messages.error(request, f"SMTP Connection Test Failed! Please check your settings.py configuration. Error Details: {str(e)}")
            return redirect('admin_dashboard')

        # 10) Instant Outbound Dispatcher (Quick message sender)
        elif "send_instant_message" in request.POST:
            recipient_type = request.POST.get('recipient_type')
            msg_type = request.POST.get('msg_type')
            message_body = request.POST.get('message_body', '').strip()
            
            if not message_body:
                messages.warning(request, "Message body cannot be empty.")
            else:
                targets = []
                if recipient_type == 'specific_student':
                    student_id = request.POST.get('specific_student')
                    student = get_object_or_404(User, pk=student_id)
                    targets.append(student)
                elif recipient_type == 'all_students':
                    targets = list(User.objects.filter(is_staff=False, is_superuser=False, is_active=True))
                
                if msg_type == 'email':
                    email_subject = request.POST.get('email_subject', 'Saral Pathshala Broadcast').strip()
                    created_count = 0
                    if recipient_type == 'manual_entry':
                        manual_email = request.POST.get('manual_email', '').strip()
                        if manual_email:
                            MailQueue.objects.create(
                                to_email=manual_email,
                                to_name="Guest/User",
                                subject=email_subject,
                                content=message_body
                            )
                            created_count = 1
                    else:
                        for student in targets:
                            MailQueue.objects.create(
                                user=student,
                                to_email=student.email,
                                to_name=student.full_name,
                                subject=email_subject,
                                content=message_body
                            )
                            created_count += 1
                    messages.success(request, f"Queued {created_count} email message(s) successfully! Go to Operations to flush the queue.")
                
                elif msg_type == 'sms':
                    created_count = 0
                    if recipient_type == 'manual_entry':
                        manual_phone = request.POST.get('manual_phone', '').strip()
                        if manual_phone:
                            SMSQueue.objects.create(
                                to_phone=manual_phone,
                                message=message_body[:160]
                            )
                            created_count = 1
                    else:
                        for student in targets:
                            if student.phone:
                                SMSQueue.objects.create(
                                    user=student,
                                    to_phone=student.phone,
                                    message=message_body[:160]
                                )
                                created_count += 1
                    messages.success(request, f"Queued {created_count} SMS message(s) successfully! Go to Operations to flush the queue.")
            return redirect('admin_dashboard')

    # 1. Basic Stats
    total_students = User.objects.filter(is_staff=False, is_superuser=False).count()
    total_courses = Course.objects.filter(is_active=True).count()
    total_exams = Exam.objects.filter(is_active=True).count()
    total_attempts_count = ExamAttempt.objects.filter(is_submitted=True).count()
    
    # 2. Real-time active students
    realtime_window = timezone.now() - timedelta(hours=3)
    realtime_attempts = ExamAttempt.objects.filter(
        is_submitted=False,
        started_at__gte=realtime_window
    ).select_related('student', 'exam')
    realtime_count = realtime_attempts.count()
    
    # 3. Exam Analysis
    exams = Exam.objects.filter(is_active=True).select_related('course')
    exam_stats = []
    
    chart_exam_titles = []
    chart_avg_scores = []
    chart_attempt_counts = []
    
    for exam in exams:
        attempts = ExamAttempt.objects.filter(exam=exam, is_submitted=True)
        count = attempts.count()
        
        # Real-time students for this specific exam
        active_now = ExamAttempt.objects.filter(
            exam=exam,
            is_submitted=False,
            started_at__gte=realtime_window
        ).count()
        
        total_marks = exam.total_marks()
        
        if count > 0:
            agg = attempts.aggregate(
                avg=Avg('score'),
                max_score=Max('score'),
                min_score=Min('score')
            )
            avg_score = round(agg['avg'] or 0.0, 2)
            max_score = round(agg['max_score'] or 0.0, 2)
            min_score = round(agg['min_score'] or 0.0, 2)
            
            # Average score percentage
            avg_pct = round((avg_score / total_marks) * 100, 1) if total_marks > 0 else 0.0
            
            # Calculate distribution (Marks pattern)
            excellent = attempts.filter(score__gte=total_marks * 0.8).count()
            good = attempts.filter(score__gte=total_marks * 0.6, score__lt=total_marks * 0.8).count()
            pass_count = attempts.filter(score__gte=total_marks * 0.4, score__lt=total_marks * 0.6).count()
            fail = attempts.filter(score__lt=total_marks * 0.4).count()
        else:
            avg_score = 0.0
            max_score = 0.0
            min_score = 0.0
            avg_pct = 0.0
            excellent = good = pass_count = fail = 0

        exam_stats.append({
            'exam': exam,
            'total_attempts': count,
            'active_now': active_now,
            'average_score': avg_score,
            'average_pct': avg_pct,
            'max_score': max_score,
            'min_score': min_score,
            'total_marks': total_marks,
            'pattern': {
                'excellent': excellent,
                'good': good,
                'pass': pass_count,
                'fail': fail
            }
        })
        
        chart_exam_titles.append(exam.title[:15] + "..." if len(exam.title) > 15 else exam.title)
        chart_avg_scores.append(avg_pct)
        chart_attempt_counts.append(count)

    # 4. Student Demographics (Educational Tier Distribution)
    plus_two_count = User.objects.filter(current_level='plus_two', is_staff=False, is_superuser=False).count()
    bachelors_count = User.objects.filter(current_level='bachelors', is_staff=False, is_superuser=False).count()
    masters_count = User.objects.filter(current_level='masters', is_staff=False, is_superuser=False).count()
    
    demographics = {
        'labels': ['+2 (XI/XII)', 'Bachelors', 'Masters'],
        'counts': [plus_two_count, bachelors_count, masters_count]
    }
    
    # 5. Course Enrollee Popularity
    chart_course_names = []
    chart_course_enrollees = []
    for c in Course.objects.filter(is_active=True):
        count = Enrollment.objects.filter(course=c).count()
        chart_course_names.append(c.name[:15] + "..." if len(c.name) > 15 else c.name)
        chart_course_enrollees.append(count)

    # 6. CRM Leads / Enrollment Requests with Pagination, Searching, and Filtering
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from django.db.models import Q

    all_courses = Course.objects.filter(is_active=True).order_by('name')
    all_exams = Exam.objects.filter(is_active=True).order_by('title')

    crm_search = request.GET.get('crm_search', '').strip()
    crm_status = request.GET.get('crm_status', '').strip()
    crm_course = request.GET.get('crm_course', '').strip()
    crm_page = request.GET.get('crm_page', '1')

    crm_qs = CourseEnrollmentRequest.objects.select_related('course').order_by('-created_at')
    if crm_search:
        crm_qs = crm_qs.filter(
            Q(name__icontains=crm_search) |
            Q(phone__icontains=crm_search) |
            Q(email__icontains=crm_search) |
            Q(message__icontains=crm_search)
        )
    if crm_status:
        crm_qs = crm_qs.filter(status=crm_status)
    if crm_course:
        crm_qs = crm_qs.filter(course_id=crm_course)

    crm_paginator = Paginator(crm_qs, 10)
    try:
        recent_requests = crm_paginator.page(crm_page)
    except PageNotAnInteger:
        recent_requests = crm_paginator.page(1)
    except EmptyPage:
        recent_requests = crm_paginator.page(crm_paginator.num_pages)

    pending_requests_count = CourseEnrollmentRequest.objects.filter(status='pending').count()
    total_requests_count = CourseEnrollmentRequest.objects.count()

    # 7. Recent Registrations & Unverified Users list with Search, Filtering & Pagination
    student_search = request.GET.get('student_search', '').strip()
    student_level = request.GET.get('student_level', '').strip()
    student_verified = request.GET.get('student_verified', '').strip()
    student_page = request.GET.get('student_page', '1')

    student_qs = User.objects.filter(is_staff=False, is_superuser=False).order_by('-date_joined')
    if student_search:
        student_qs = student_qs.filter(
            Q(full_name__icontains=student_search) |
            Q(email__icontains=student_search) |
            Q(phone__icontains=student_search) |
            Q(previous_institute__icontains=student_search)
        )
    if student_level:
        student_qs = student_qs.filter(current_level=student_level)
    if student_verified == 'verified':
        student_qs = student_qs.filter(is_phone_verified=True)
    elif student_verified == 'unverified':
        student_qs = student_qs.filter(is_phone_verified=False)

    student_paginator = Paginator(student_qs, 10)
    try:
        recent_students = student_paginator.page(student_page)
    except PageNotAnInteger:
        recent_students = student_paginator.page(1)
    except EmptyPage:
        recent_students = student_paginator.page(student_paginator.num_pages)

    unverified_users = User.objects.filter(is_email_verified=False, is_staff=False, is_superuser=False).order_by('-date_joined')[:20]

    # 8. Queue Analytics & Failure Monitoring
    recent_failed_emails = MailQueue.objects.filter(status=QueueStatus.FAILED).order_by('-created_at')[:10]
    recent_pending_emails = MailQueue.objects.filter(status=QueueStatus.PENDING).order_by('-created_at')[:10]
    mail_queue_stats = {
        'pending': MailQueue.objects.filter(status=QueueStatus.PENDING).count(),
        'sent': MailQueue.objects.filter(status=QueueStatus.SENT).count(),
        'failed': MailQueue.objects.filter(status=QueueStatus.FAILED).count(),
        'total': MailQueue.objects.count()
    }

    recent_failed_sms = SMSQueue.objects.filter(status=QueueStatus.FAILED).order_by('-created_at')[:10]
    recent_pending_sms = SMSQueue.objects.filter(status=QueueStatus.PENDING).order_by('-created_at')[:10]
    sms_queue_stats = {
        'pending': SMSQueue.objects.filter(status=QueueStatus.PENDING).count(),
        'sent': SMSQueue.objects.filter(status=QueueStatus.SENT).count(),
        'failed': SMSQueue.objects.filter(status=QueueStatus.FAILED).count(),
        'total': SMSQueue.objects.count()
    }

    # 9. Comprehensive Recent Exam Attempts for Retake / Review management with Search, Filters & Pagination
    attempt_search = request.GET.get('attempt_search', '').strip()
    attempt_exam = request.GET.get('attempt_exam', '').strip()
    attempt_status = request.GET.get('attempt_status', '').strip()
    attempt_page = request.GET.get('attempt_page', '1')

    attempts_qs = ExamAttempt.objects.select_related('student', 'exam', 'exam__course').order_by('-started_at')
    if attempt_search:
        attempts_qs = attempts_qs.filter(
            Q(student__full_name__icontains=attempt_search) |
            Q(student__email__icontains=attempt_search) |
            Q(student__phone__icontains=attempt_search)
        )
    if attempt_exam:
        attempts_qs = attempts_qs.filter(exam_id=attempt_exam)
    if attempt_status == 'submitted':
        attempts_qs = attempts_qs.filter(is_submitted=True)
    elif attempt_status == 'active':
        attempts_qs = attempts_qs.filter(is_submitted=False)

    attempts_paginator = Paginator(attempts_qs, 10)
    try:
        recent_attempts = attempts_paginator.page(attempt_page)
    except PageNotAnInteger:
        recent_attempts = attempts_paginator.page(1)
    except EmptyPage:
        recent_attempts = attempts_paginator.page(attempts_paginator.num_pages)

    # CRM Conversion Analytics
    approved_leads_count = CourseEnrollmentRequest.objects.filter(status='approved').count()
    lead_conversion_rate = round((approved_leads_count / total_requests_count) * 100, 1) if total_requests_count > 0 else 0.0
    
    # Student Verification rate
    verified_students_count = User.objects.filter(is_phone_verified=True, is_staff=False, is_superuser=False).count()
    student_verification_rate = round((verified_students_count / total_students) * 100, 1) if total_students > 0 else 0.0
    
    # Mail & SMS Delivery Rates
    total_mail_count = MailQueue.objects.count()
    sent_mail_count = MailQueue.objects.filter(status=QueueStatus.SENT).count()
    mail_delivery_rate = round((sent_mail_count / total_mail_count) * 100, 1) if total_mail_count > 0 else 100.0
    
    total_sms_count = SMSQueue.objects.count()
    sent_sms_count = SMSQueue.objects.filter(status=QueueStatus.SENT).count()
    sms_delivery_rate = round((sent_sms_count / total_sms_count) * 100, 1) if total_sms_count > 0 else 100.0

    # Platform-wide or exam-specific Leaderboard / Top Performers
    leaderboard_exam_id = request.GET.get('leaderboard_exam', '').strip()
    from django.db.models import Avg, Max, Count
    leaderboard_attempts = ExamAttempt.objects.filter(is_submitted=True)
    if leaderboard_exam_id:
        leaderboard_attempts = leaderboard_attempts.filter(exam_id=leaderboard_exam_id)
        
    top_performers = leaderboard_attempts\
        .values('student__full_name', 'student__email')\
        .annotate(avg_score=Avg('score'), max_score=Max('score'), attempts_count=Count('id'))\
        .order_by('-avg_score')[:10]
        
    # All students query for Quick Dispatcher select box
    all_students_list = User.objects.filter(is_staff=False, is_superuser=False).order_by('full_name')

    context = {
        'total_students': total_students,
        'total_courses': total_courses,
        'total_exams': total_exams,
        'total_attempts_count': total_attempts_count,
        'realtime_count': realtime_count,
        'realtime_attempts': realtime_attempts,
        'exam_stats': exam_stats,
        'recent_requests': recent_requests,
        'pending_requests_count': pending_requests_count,
        'total_requests_count': total_requests_count,
        'recent_students': recent_students,
        'unverified_users': unverified_users,
        
        # Advanced Analytics & Broadcast
        'lead_conversion_rate': lead_conversion_rate,
        'student_verification_rate': student_verification_rate,
        'mail_delivery_rate': mail_delivery_rate,
        'sms_delivery_rate': sms_delivery_rate,
        'top_performers': top_performers,
        'all_students_list': all_students_list,
        
        # Filter choices
        'all_courses': all_courses,
        'all_exams': all_exams,
        
        # Filter values
        'crm_search': crm_search,
        'crm_status': crm_status,
        'crm_course_id': int(crm_course) if crm_course.isdigit() else '',
        'student_search': student_search,
        'student_level': student_level,
        'student_verified': student_verified,
        'leaderboard_exam_id': int(leaderboard_exam_id) if leaderboard_exam_id.isdigit() else '',
        'attempt_search': attempt_search,
        'attempt_exam_id': int(attempt_exam) if attempt_exam.isdigit() else '',
        'attempt_status': attempt_status,
        
        # Queues Context
        'recent_failed_emails': recent_failed_emails,
        'recent_pending_emails': recent_pending_emails,
        'mail_queue_stats': mail_queue_stats,
        'recent_failed_sms': recent_failed_sms,
        'recent_pending_sms': recent_pending_sms,
        'sms_queue_stats': sms_queue_stats,
        
        # Attempt Reviews
        'recent_attempts': recent_attempts,
        
        # Charts JSON
        'chart_exam_titles_json': json.dumps(chart_exam_titles),
        'chart_avg_scores_json': json.dumps(chart_avg_scores),
        'chart_attempt_counts_json': json.dumps(chart_attempt_counts),
        
        'demographics_json': json.dumps(demographics),
        'chart_course_names_json': json.dumps(chart_course_names),
        'chart_course_enrollees_json': json.dumps(chart_course_enrollees),
    }
    
    return render(request, 'admin_custom/dashboard.html', context)


@login_required
def admin_manage_enrollment_request(request, request_id, action):
    """
    Staff / Superuser CRM utility to approve, reject, or mark guest/student leads as contacted.
    If approved, automatically links/enrolls user to the course track if a student account exists.
    """
    if not (request.user.is_staff or request.user.is_superuser):
        raise Http404("You do not have permission to view this page.")

    enroll_req = get_object_or_404(CourseEnrollmentRequest, pk=request_id)
    
    if action == 'approve':
        enroll_req.status = 'approved'
        enroll_req.save()
        
        # Auto-create user enrollment if the user exists in database
        student = User.objects.filter(email=enroll_req.email).first()
        if not student and enroll_req.phone:
            student = User.objects.filter(phone=enroll_req.phone).first()
            
        if student:
            # Check for existing enrollment
            already_enrolled = Enrollment.objects.filter(user=student, course=enroll_req.course).exists()
            if not already_enrolled:
                Enrollment.objects.create(
                    user=student,
                    course=enroll_req.course,
                    enrolled_by=request.user
                )
                messages.success(request, f"Enrollment lead approved! Student '{student.get_full_name()}' has been enrolled in '{enroll_req.course.name}' successfully.")
            else:
                messages.success(request, f"Enrollment lead approved! Student was already enrolled.")
        else:
            messages.warning(request, f"Request approved in CRM, but no user account could be located for email '{enroll_req.email}' or phone '{enroll_req.phone}'. Please register the student first.")
            
    elif action == 'contact':
        enroll_req.status = 'contacted'
        enroll_req.save()
        messages.info(request, f"Enrollment lead status updated to Contacted / Followed Up.")
        
    elif action == 'reject':
        enroll_req.status = 'rejected'
        enroll_req.save()
        messages.error(request, f"Enrollment lead rejected.")
        
    next_url = request.META.get('HTTP_REFERER') or reverse('admin_dashboard')
    return redirect(next_url)


@login_required
def admin_export_exam_results(request, exam_id):
    """
    Export all submitted exam attempts for a specific exam as a detailed Excel workbook.
    Supports exporting Overview, Section-wise, and Question-wise sheets selectively or all together.
    """
    if not (request.user.is_staff or request.user.is_superuser):
        raise Http404("You do not have permission to perform this action.")

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    exam = get_object_or_404(Exam, pk=exam_id)
    attempts = ExamAttempt.objects.filter(exam=exam, is_submitted=True).select_related('student').order_by('-score')
    sections = list(exam.sections.all().order_by('order', 'id'))
    questions = list(exam.questions.all().order_by('section__order', 'order', 'id'))
    
    export_type = request.GET.get('type', 'all').lower()
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    # Define beautiful styles
    title_font = Font(name='Segoe UI', size=16, bold=True, color='1E293B')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    subheader_font = Font(name='Segoe UI', size=11, bold=True, color='1E293B')
    data_font = Font(name='Segoe UI', size=10, color='334155')
    bold_data_font = Font(name='Segoe UI', size=10, bold=True, color='1E293B')
    
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    sec_header_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
    zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # ── SHEET 1: SUMMARY RESULTS (Overview) ──────────────────────────────────
    if export_type in ['all', 'overview']:
        ws1 = wb.create_sheet(title="Summary Results")
        ws1.views.sheetView[0].showGridLines = True
        
        # Title Block
        ws1.append([])
        ws1.cell(row=2, column=2, value=f"Exam Results Summary: {exam.title}").font = title_font
        ws1.cell(row=3, column=2, value=f"Course: {exam.course.name} | Total Marks: {exam.total_marks()}").font = Font(name='Segoe UI', size=11, italic=True, color='64748B')
        ws1.append([])
        ws1.append([])
        
        headers1 = [
            'Rank', 'Student Name', 'Student Email', 'Student Phone', 
            'Score Obtained', 'Total Marks', 'Correct Answers', 
            'Wrong Answers', 'Unattempted Questions', 'Accuracy %', 
            'Started At', 'Submitted At'
        ]
        ws1.append(headers1)
        
        # Format Headers row
        header_row_idx = 6
        for col_idx in range(1, len(headers1) + 1):
            cell = ws1.cell(row=header_row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        total_marks = exam.total_marks()
        for rank, att in enumerate(attempts, 1):
            pct = round((att.score / total_marks) * 100, 1) if total_marks > 0 else 0.0
            row_data = [
                rank,
                att.student.full_name,
                att.student.email,
                att.student.phone or 'N/A',
                att.score,
                total_marks,
                att.correct_count,
                att.wrong_count,
                att.unattempted_count,
                pct / 100,  # Excel uses decimal for percentage formatting
                att.started_at.strftime('%Y-%m-%d %H:%M:%S'),
                att.completed_at.strftime('%Y-%m-%d %H:%M:%S') if att.completed_at else 'N/A'
            ]
            ws1.append(row_data)
            curr_row = ws1.max_row
            
            # Style row cells
            is_even = (rank % 2 == 0)
            for col_idx in range(1, len(headers1) + 1):
                cell = ws1.cell(row=curr_row, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                if is_even:
                    cell.fill = zebra_fill
                
                # Alignments & Formats
                if col_idx in [1, 5, 6, 7, 8, 9]:
                    cell.alignment = Alignment(horizontal='center')
                elif col_idx in [10]:
                    cell.alignment = Alignment(horizontal='center')
                    cell.number_format = '0.0%'
                elif col_idx in [11, 12]:
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')

        # Auto-fit columns for Sheet 1
        for col in ws1.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            # Skip title block rows when calculating width
            for cell in col:
                if cell.row > 4 and cell.value:
                    val_str = str(cell.value)
                    if cell.number_format == '0.0%' and isinstance(cell.value, float):
                        val_str = f"{cell.value * 100:.1f}%"
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # ── SHEET 2: SECTION-WISE ANALYSIS ───────────────────────────────────────
    if export_type in ['all', 'section']:
        ws2 = wb.create_sheet(title="Section-wise Analysis")
        ws2.views.sheetView[0].showGridLines = True
        
        # Title Block
        ws2.append([])
        ws2.cell(row=2, column=2, value=f"Section-wise Marks Analysis: {exam.title}").font = title_font
        ws2.append([])
        ws2.append([])
        
        # Headers
        # Row 5 will contain Section Group titles (merged columns)
        # Row 6 will contain detail column headers
        headers_row5 = ['', '', '', '']  # Rank, Name, Email, Phone
        headers_row6 = ['Rank', 'Student Name', 'Student Email', 'Student Phone']
        
        # For each section, add a 4-column group header
        for sec in sections:
            headers_row5.extend([sec.title, '', '', ''])
            headers_row6.extend(['Score', 'Correct', 'Wrong', 'Unattempted'])
        
        # Overall summary columns
        headers_row5.extend(['Overall Total', '', ''])
        headers_row6.extend(['Total Score', 'Total Correct', 'Total Wrong'])
        
        ws2.append(headers_row5)
        ws2.append(headers_row6)
        
        row5_idx = 5
        row6_idx = 6
        
        # Merge cells for sections group headers on Row 5
        # Standard columns are 1 to 4
        col_idx = 5
        for sec in sections:
            ws2.merge_cells(start_row=row5_idx, start_column=col_idx, end_row=row5_idx, end_column=col_idx+3)
            # Style the merged group cell
            group_cell = ws2.cell(row=row5_idx, column=col_idx)
            group_cell.font = header_font
            group_cell.fill = header_fill
            group_cell.alignment = Alignment(horizontal='center', vertical='center')
            group_cell.border = thin_border
            
            # Apply border to all merged cells
            for c in range(col_idx, col_idx + 4):
                ws2.cell(row=row5_idx, column=c).border = thin_border
                ws2.cell(row=row5_idx, column=c).fill = header_fill
            col_idx += 4
            
        # Merge for Overall Summary Group Header
        ws2.merge_cells(start_row=row5_idx, start_column=col_idx, end_row=row5_idx, end_column=col_idx+2)
        group_cell = ws2.cell(row=row5_idx, column=col_idx)
        group_cell.font = header_font
        group_cell.fill = header_fill
        group_cell.alignment = Alignment(horizontal='center', vertical='center')
        for c in range(col_idx, col_idx + 3):
            ws2.cell(row=row5_idx, column=c).border = thin_border
            ws2.cell(row=row5_idx, column=c).fill = header_fill
            
        # Merge first 4 columns vertically for header
        for c in range(1, 5):
            ws2.merge_cells(start_row=row5_idx, start_column=c, end_row=row6_idx, end_column=c)
            v_cell = ws2.cell(row=row5_idx, column=c)
            v_cell.font = header_font
            v_cell.fill = header_fill
            v_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws2.cell(row=row6_idx, column=c).border = thin_border
            v_cell.border = thin_border
            
        # Format subheaders (Row 6 details)
        for c in range(5, col_idx + 3):
            cell = ws2.cell(row=row6_idx, column=c)
            cell.font = subheader_font
            cell.fill = sec_header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Fill data
        for rank, att in enumerate(attempts, 1):
            # We need a map of this student's selected options
            student_answers = {qa.question_id: qa.selected_option for qa in att.question_attempts.all()}
            
            row_data = [
                rank,
                att.student.full_name,
                att.student.email,
                att.student.phone or 'N/A'
            ]
            
            # Calculate section-wise breakdown
            for sec in sections:
                sec_questions = [q for q in questions if q.section_id == sec.id]
                sec_score = 0.0
                sec_correct = 0
                sec_wrong = 0
                sec_unattempted = 0
                
                for q in sec_questions:
                    selected = student_answers.get(q.id)
                    if not selected:
                        sec_unattempted += 1
                    elif selected == q.correct_option:
                        sec_score += q.get_correct_marks()
                        sec_correct += 1
                    else:
                        sec_score -= q.get_negative_marks()
                        sec_wrong += 1
                row_data.extend([round(sec_score, 4), sec_correct, sec_wrong, sec_unattempted])
                
            row_data.extend([att.score, att.correct_count, att.wrong_count])
            ws2.append(row_data)
            curr_row = ws2.max_row
            
            is_even = (rank % 2 == 0)
            # Style rows
            for c in range(1, len(row_data) + 1):
                cell = ws2.cell(row=curr_row, column=c)
                cell.font = data_font
                cell.border = thin_border
                if is_even:
                    cell.fill = zebra_fill
                if c <= 4:
                    cell.alignment = Alignment(horizontal='left' if c > 1 else 'center')
                else:
                    cell.alignment = Alignment(horizontal='center')
                    # highlight total score
                    if c == len(row_data) - 2:
                        cell.font = bold_data_font

        # Set column widths for Sheet 2
        for col in ws2.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row > 4 and cell.value:
                    val_str = str(cell.value)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            ws2.column_dimensions[col_letter].width = max(max_len + 4, 11)

    # ── SHEET 3: QUESTION-WISE DETAIL ────────────────────────────────────────
    if export_type in ['all', 'question']:
        ws3 = wb.create_sheet(title="Question-wise Detail")
        ws3.views.sheetView[0].showGridLines = True
        
        ws3.append([])
        ws3.cell(row=2, column=2, value=f"Question-wise Selected Options: {exam.title}").font = title_font
        ws3.append([])
        ws3.append([])
        
        # Headers
        # Row 5: Section titles
        # Row 6: Q numbers + Correct option
        q_row5 = ['', '', '', '']
        q_row6 = ['Rank', 'Student Name', 'Student Email', 'Student Phone']
        
        for idx, q in enumerate(questions, 1):
            q_row5.append(q.section.title)
            q_row6.append(f"Q{idx} (Ans: {q.correct_option or 'N/A'})")
            
        ws3.append(q_row5)
        ws3.append(q_row6)
        
        # Merge and style headers
        for c in range(1, 5):
            ws3.merge_cells(start_row=5, start_column=c, end_row=6, end_column=c)
            v_cell = ws3.cell(row=5, column=c)
            v_cell.font = header_font
            v_cell.fill = header_fill
            v_cell.alignment = Alignment(horizontal='center', vertical='center')
            v_cell.border = thin_border
            ws3.cell(row=6, column=c).border = thin_border
            
        # Group sections header
        current_sec_title = None
        start_col = 5
        for idx, q in enumerate(questions, 5):
            # Format Q headers on Row 6
            cell_q = ws3.cell(row=6, column=idx)
            cell_q.font = subheader_font
            cell_q.fill = sec_header_fill
            cell_q.alignment = Alignment(horizontal='center', vertical='center')
            cell_q.border = thin_border
            
            # Merge section header on Row 5 if it's the same section
            sec_title = q.section.title
            if current_sec_title is None:
                current_sec_title = sec_title
                start_col = idx
            elif sec_title != current_sec_title:
                # Merge previous
                ws3.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=idx-1)
                grp_cell = ws3.cell(row=5, column=start_col)
                grp_cell.font = header_font
                grp_cell.fill = header_fill
                grp_cell.alignment = Alignment(horizontal='center', vertical='center')
                for c in range(start_col, idx):
                    ws3.cell(row=5, column=c).border = thin_border
                    ws3.cell(row=5, column=c).fill = header_fill
                
                # Start new
                current_sec_title = sec_title
                start_col = idx
                
        # Merge the last group
        if current_sec_title is not None and start_col < len(questions) + 5:
            end_col = len(questions) + 4
            ws3.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col)
            grp_cell = ws3.cell(row=5, column=start_col)
            grp_cell.font = header_font
            grp_cell.fill = header_fill
            grp_cell.alignment = Alignment(horizontal='center', vertical='center')
            for c in range(start_col, end_col + 1):
                ws3.cell(row=5, column=c).border = thin_border
                ws3.cell(row=5, column=c).fill = header_fill

        # Fill student question-by-question options
        correct_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')  # light green
        wrong_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')    # light red
        unattempted_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid') # light slate
        
        for rank, att in enumerate(attempts, 1):
            student_answers = {qa.question_id: qa.selected_option for qa in att.question_attempts.all()}
            
            row_data = [
                rank,
                att.student.full_name,
                att.student.email,
                att.student.phone or 'N/A'
            ]
            
            for q in questions:
                selected = student_answers.get(q.id)
                if not selected:
                    row_data.append("Unattempted")
                else:
                    status = "Correct" if selected == q.correct_option else "Wrong"
                    row_data.append(f"Opt {selected} ({status})")
                    
            ws3.append(row_data)
            curr_row = ws3.max_row
            
            is_even = (rank % 2 == 0)
            # Style rows
            for c in range(1, len(row_data) + 1):
                cell = ws3.cell(row=curr_row, column=c)
                cell.font = data_font
                cell.border = thin_border
                if is_even:
                    cell.fill = zebra_fill
                if c <= 4:
                    cell.alignment = Alignment(horizontal='left' if c > 1 else 'center')
                else:
                    cell.alignment = Alignment(horizontal='center')
                    val = str(cell.value)
                    if "Correct" in val:
                        cell.fill = correct_fill
                        cell.font = Font(name='Segoe UI', size=10, color='1B5E20', bold=True)
                    elif "Wrong" in val:
                        cell.fill = wrong_fill
                        cell.font = Font(name='Segoe UI', size=10, color='B71C1C')
                    elif "Unattempted" in val:
                        cell.fill = unattempted_fill
                        cell.font = Font(name='Segoe UI', size=10, color='64748B')

        # Set column widths for Sheet 3
        for col in ws3.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row > 4 and cell.value:
                    val_str = str(cell.value)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            ws3.column_dimensions[col_letter].width = max(max_len + 4, 11)

    # Prepare HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    if export_type == 'overview':
        filename = f"results_overview_{exam.slug}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    elif export_type == 'section':
        filename = f"results_sectionwise_{exam.slug}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    elif export_type == 'question':
        filename = f"results_questionwise_{exam.slug}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    else:
        filename = f"results_detailed_{exam.slug}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def admin_api_get_sections_paragraphs(request):
    """
    JSON API for dynamic dropdown autopopulation in Django Admin.
    Given an exam_id, returns lists of its sections and paragraphs.
    """
    if not (request.user.is_staff or request.user.is_superuser):
        raise Http404("You do not have permission to perform this action.")

    from django.http import JsonResponse
    from apps.exam.models import Section, Paragraph

    exam_id = request.GET.get('exam_id')
    sections = []
    paragraphs = []
    if exam_id:
        sections = list(Section.objects.filter(exam_id=exam_id).order_by('order', 'id').values('id', 'title'))
        paragraphs = list(Paragraph.objects.filter(section__exam_id=exam_id).order_by('order', 'id').values('id', 'title', 'section_id'))

    return JsonResponse({
        'sections': sections,
        'paragraphs': paragraphs
    })


# ── Exam CRUD Views ─────────────────────────────────────────────────────────

@login_required
def admin_manage_exam(request, exam_id=None):
    """
    Sleek, custom page to create or update mock exams.
    Inherits TinyMCE rich editor and standard validations.
    """
    if not (request.user.is_staff or request.user.is_superuser):
        raise Http404("You do not have permission to view this page.")
        
    from apps.exam.models import Exam
    from apps.exam.forms import ExamAdminForm
    from apps.pages.models import Course
    from apps.exam.utils import invalidate_exam_cache
    
    exam = None
    if exam_id:
        exam = get_object_or_404(Exam, pk=exam_id)
        
    if request.method == 'POST':
        form = ExamAdminForm(request.POST, instance=exam)
        if form.is_valid():
            saved_exam = form.save()
            invalidate_exam_cache(saved_exam.pk)
            messages.success(request, f"Mock exam '{saved_exam.title}' saved successfully!")
            return redirect('admin_dashboard')
        else:
            messages.error(request, "Failed to save. Please correct form errors.")
    else:
        form = ExamAdminForm(instance=exam)
        
    return render(request, 'admin_custom/manage_exam.html', {
        'form': form,
        'exam': exam,
        'courses': Course.objects.all()
    })


@login_required
def admin_delete_exam(request, exam_id):
    """
    Purge a mock exam, sections, questions and attempts with cache invalidation.
    """
    if not (request.user.is_staff or request.user.is_superuser):
        raise Http404("You do not have permission to perform this action.")
        
    from apps.exam.models import Exam
    from apps.exam.utils import invalidate_exam_cache
    
    exam = get_object_or_404(Exam, pk=exam_id)
    title = exam.title
    invalidate_exam_cache(exam.pk)
    exam.delete()
    messages.error(request, f"Mock exam '{title}' and all its contents were deleted.")
    return redirect('admin_dashboard')


# ── Section CRUD Views ──────────────────────────────────────────────────────

@login_required
def admin_manage_sections(request, exam_id):
    """
    CRUD management for sections of a given mock exam.
    Supports display ordering, override defaults, and custom scoring marks.
    """
    if not (request.user.is_staff or request.user.is_superuser):
        raise Http404("You do not have permission to view this page.")
        
    from apps.exam.models import Exam, Section
    from django import forms
    from apps.exam.utils import invalidate_exam_cache
    
    exam = get_object_or_404(Exam, pk=exam_id)
    sections = exam.sections.all().order_by('order', 'id')
    
    class SectionForm(forms.ModelForm):
        class Meta:
            model = Section
            fields = ['title', 'description', 'order', 'override_scoring',
                      'custom_correct_marks', 'custom_negative_marks', 'custom_has_negative']
            widgets = {
                'title': forms.TextInput(attrs={'class': 'form-control', 'required': 'true'}),
                'description': forms.Textarea(attrs={'class': 'form-control', 'rows': 2}),
                'order': forms.NumberInput(attrs={'class': 'form-control'}),
                'override_scoring': forms.CheckboxInput(attrs={'class': 'form-check-input', 'id': 'id_override_scoring'}),
                'custom_correct_marks': forms.NumberInput(attrs={'class': 'form-control', 'step': 'any'}),
                'custom_negative_marks': forms.NumberInput(attrs={'class': 'form-control', 'step': 'any'}),
                'custom_has_negative': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            }
            
    section_id = request.GET.get('edit_section')
    edit_section = None
    if section_id:
        edit_section = get_object_or_404(Section, pk=section_id, exam=exam)
        
    if request.method == 'POST':
        form = SectionForm(request.POST, instance=edit_section)
        if form.is_valid():
            sec = form.save(commit=False)
            sec.exam = exam
            sec.save()
            invalidate_exam_cache(exam.pk)
            messages.success(request, f"Section '{sec.title}' saved successfully!")
            return redirect(reverse('admin_manage_sections', kwargs={'exam_id': exam.pk}))
        else:
            messages.error(request, "Failed to save section. Verify fields.")
    else:
        form = SectionForm(instance=edit_section)
        
    return render(request, 'admin_custom/manage_sections.html', {
        'exam': exam,
        'sections': sections,
        'form': form,
        'edit_section': edit_section
    })


@login_required
def admin_delete_section(request, exam_id, section_id):
    """
    Delete section and invalidate exam cache.
    """
    if not (request.user.is_staff or request.user.is_superuser):
        raise Http404("You do not have permission to perform this action.")
        
    from apps.exam.models import Section
    from apps.exam.utils import invalidate_exam_cache
    
    section = get_object_or_404(Section, pk=section_id, exam_id=exam_id)
    title = section.title
    section.delete()
    invalidate_exam_cache(exam_id)
    messages.error(request, f"Section '{title}' deleted.")
    return redirect(reverse('admin_manage_sections', kwargs={'exam_id': exam_id}))


# ── Paragraph CRUD Views ────────────────────────────────────────────────────

@login_required
def admin_manage_paragraphs(request, section_id):
    """
    CRUD management for reading passages / paragraphs of a section.
    """
    if not (request.user.is_staff or request.user.is_superuser):
        raise Http404("You do not have permission to view this page.")
        
    from apps.exam.models import Section, Paragraph
    from apps.exam.forms import ParagraphAdminForm
    from apps.exam.utils import invalidate_exam_cache
    
    section = get_object_or_404(Section, pk=section_id)
    paragraphs = section.paragraphs.all().order_by('order', 'id')
    
    paragraph_id = request.GET.get('edit_paragraph')
    edit_paragraph = None
    if paragraph_id:
        edit_paragraph = get_object_or_404(Paragraph, pk=paragraph_id, section=section)
        
    if request.method == 'POST':
        form = ParagraphAdminForm(request.POST, instance=edit_paragraph)
        if form.is_valid():
            para = form.save(commit=False)
            para.section = section
            para.save()
            invalidate_exam_cache(section.exam_id)
            messages.success(request, f"Passage '{para.title or para.id}' saved successfully!")
            return redirect(reverse('admin_manage_paragraphs', kwargs={'section_id': section.pk}))
        else:
            messages.error(request, "Failed to save passage.")
    else:
        form = ParagraphAdminForm(instance=edit_paragraph)
        
    return render(request, 'admin_custom/manage_paragraphs.html', {
        'section': section,
        'paragraphs': paragraphs,
        'form': form,
        'edit_paragraph': edit_paragraph
    })


@login_required
def admin_delete_paragraph(request, section_id, paragraph_id):
    """
    Delete a passage and invalidate cache.
    """
    if not (request.user.is_staff or request.user.is_superuser):
        raise Http404("You do not have permission to perform this action.")
        
    from apps.exam.models import Paragraph
    from apps.exam.utils import invalidate_exam_cache
    
    paragraph = get_object_or_404(Paragraph, pk=paragraph_id, section_id=section_id)
    title = paragraph.title or f"Passage #{paragraph.id}"
    exam_id = paragraph.section.exam_id
    paragraph.delete()
    invalidate_exam_cache(exam_id)
    messages.error(request, f"Passage '{title}' deleted.")
    return redirect(reverse('admin_manage_paragraphs', kwargs={'section_id': section_id}))


@login_required
def admin_api_get_question_difficulty(request):
    """
    JSON API for dynamic question difficulty breakdown in Mock Exam Analytics.
    """
    if not (request.user.is_staff or request.user.is_superuser):
        raise Http404("You do not have permission to perform this action.")
        
    from django.http import JsonResponse
    from apps.exam.models import Exam, QuestionAttempt
    from django.db.models import F, Count
    
    exam_id = request.GET.get('exam_id')
    if not exam_id:
        return JsonResponse({'error': 'Missing exam_id'}, status=400)
        
    exam = get_object_or_404(Exam, pk=exam_id)
    questions = list(exam.questions.all().order_by('section__order', 'order', 'id'))
    total_attempts = exam.attempts.filter(is_submitted=True).count()
    
    stats = []
    if total_attempts > 0:
        q_attempts = QuestionAttempt.objects.filter(exam_attempt__exam=exam, exam_attempt__is_submitted=True)
        
        correct_counts = {
            qa['question_id']: qa['count']
            for qa in q_attempts.filter(selected_option=F('question__correct_option'))
            .values('question_id')
            .annotate(count=Count('id'))
        }
        
        answered_counts = {
            qa['question_id']: qa['count']
            for qa in q_attempts.exclude(selected_option__isnull=True).exclude(selected_option='')
            .values('question_id')
            .annotate(count=Count('id'))
        }
        
        import re
        for idx, q in enumerate(questions, 1):
            correct = correct_counts.get(q.id, 0)
            answered = answered_counts.get(q.id, 0)
            unattempted = total_attempts - answered
            wrong = answered - correct
            success_rate = round((correct / total_attempts) * 100, 1)
            
            if success_rate < 30:
                difficulty = "Hard"
                difficulty_color = "danger"
            elif success_rate > 70:
                difficulty = "Easy"
                difficulty_color = "success"
            else:
                difficulty = "Medium"
                difficulty_color = "warning"
                
            clean_text = re.sub(r'<[^>]+>', '', q.question_text)
            clean_text = clean_text[:90] + '...' if len(clean_text) > 90 else clean_text
            
            stats.append({
                'num': idx,
                'text': clean_text,
                'correct': correct,
                'wrong': wrong,
                'unattempted': unattempted,
                'success_rate': success_rate,
                'difficulty': difficulty,
                'difficulty_color': difficulty_color
            })
            
    # Calculate score ranges distribution
    excellent_count = 0
    good_count = 0
    pass_count = 0
    fail_count = 0
    if total_attempts > 0:
        attempts_exam = exam.attempts.filter(is_submitted=True)
        total_marks = exam.total_marks()
        if total_marks > 0:
            excellent_count = attempts_exam.filter(score__gte=total_marks * 0.8).count()
            good_count = attempts_exam.filter(score__gte=total_marks * 0.6, score__lt=total_marks * 0.8).count()
            pass_count = attempts_exam.filter(score__gte=total_marks * 0.4, score__lt=total_marks * 0.6).count()
            fail_count = attempts_exam.filter(score__lt=total_marks * 0.4).count()
            
    return JsonResponse({
        'total_attempts': total_attempts,
        'stats': stats,
        'distribution': {
            'fail': fail_count,
            'pass': pass_count,
            'good': good_count,
            'excellent': excellent_count
        }
    })